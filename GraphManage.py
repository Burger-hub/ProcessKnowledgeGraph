# coding:utf-8
import py2neo
from py2neo import NodeMatcher,RelationshipMatcher
from openpyxl import load_workbook

# 连接neo4j数据库，输入地址、用户名、密码
graph = py2neo.Graph('bolt://localhost:7687', auth=('neo4j', '1'))


matcher = NodeMatcher(graph)
rmatcher = RelationshipMatcher(graph)
def findNode(label,Name):
    node = matcher.match(label).where(name=Name).first()   #取出符合条件的第一个节点
    #node = list(matcher.match(label).where(age=35)) #取出符合条件的节点列表
    return node

def createNode(label, **kwargs): 
    if (findNode(label, kwargs['name'])==None):
        f= py2neo.Node(label, **kwargs)
        #print(findNode(label, kwargs['name']))
        graph.create(f)
        return f
    else:
        print("[{}] node has already existed".format(kwargs['name']))

def creatRelat(Node1,relat,Node2):
    r= py2neo.Relationship(Node1, relat, Node2)
    #print(r)
    graph.create(r)
    return r

def addProperties(label,name):##unfinished
    node=findNode(label,name)

def xlsx2nodes():
    workbook = load_workbook(".\data.xlsx")
    #print(workbook.sheetnames)
    nodeSheet = workbook['nodeData']
    #特征节点
    max_row_feature = max((bb.row for bb in nodeSheet['B'] if bb.value))
    #print(max_row_feature)
    for i in range(2, max_row_feature+1):
        featureName=nodeSheet['B{}'.format(i)].value
        createNode('feature',name= featureName )
        print("featureNode [{}] creat".format(featureName))
    print("------created {} featureNode------ ".format(max_row_feature-1))    
    #工艺节点
    max_row_process = max((bb.row for bb in nodeSheet['D'] if bb.value))
    for i in range(2, max_row_process+1):
        pName=nodeSheet['D{}'.format(i)].value
        pLowerIT=nodeSheet['E{}'.format(i)].value
        pUpperIT=nodeSheet['F{}'.format(i)].value
        pLowerRa=nodeSheet['G{}'.format(i)].value
        pUpperRa=nodeSheet['H{}'.format(i)].value
        #print(processName,processIT,processRa)
        #print(type(processIT)) #目前均以字符串的形式存入节点
        createNode('process',name= pName,lowerIT=pLowerIT,upperIT=pUpperIT,lowerRa=pLowerRa,upperRa=pUpperRa )
        print("processNode [{}] creat".format(pName))
    print("------created {} processNode------ ".format(max_row_process-1))
    #设备节点
    max_row_equipment = max((bb.row for bb in nodeSheet['J'] if bb.value))
    for i in range(2, max_row_equipment+1):
        eName=nodeSheet['J{}'.format(i)].value
        eLowerSize=nodeSheet['K{}'.format(i)].value
        eUpperSize=nodeSheet['L{}'.format(i)].value
        eLowerIT=nodeSheet['M{}'.format(i)].value
        eUpperIT=nodeSheet['N{}'.format(i)].value
        eNotes=nodeSheet['O{}'.format(i)].value
        #print(processName,processIT,processRa)
        #print(type(processIT)) #目前均以字符串的形式存入节点
        createNode('equipment',name= eName,lowerSize=eLowerSize,upperSize=eUpperSize,lowerIT=eLowerIT,upperIT=eUpperIT,notes=eNotes )
        print("equipmentNode [{}] creat".format(pName))
    print("------created {} equipmentNode------ ".format(max_row_equipment-1))
    #刀具节点
    max_row_tool = max((bb.row for bb in nodeSheet['Q'] if bb.value))
    for i in range(2, max_row_tool+1):
        tName=nodeSheet['Q{}'.format(i)].value
        createNode('tool',name= tName)
        print("tool [{}] creat".format(tName))
    print("------created {} equipmentNode------ ".format(max_row_tool-1))

def xlsx2relats():
    workbook = load_workbook(".\data.xlsx")
    #print(workbook.sheetnames)
    nodeSheet = workbook['relatData']
    #可用工艺
    max_row_use = max((bb.row for bb in nodeSheet['B'] if bb.value))
    #print(max_row_feature)
    for i in range(2, max_row_use+1):
        relatType=nodeSheet['B{}'.format(i)].value
        startNode=findNode('feature',nodeSheet['A{}'.format(i)].value)
        endNode=findNode('process',nodeSheet['C{}'.format(i)].value)
        r1=creatRelat(startNode,relatType,endNode)
        print("{} creat".format(r1))
    print("------created {} '可用工艺' relation------ ".format(max_row_use-1)) 
    #后序工艺
    max_row_pre = max((bb.row for bb in nodeSheet['E'] if bb.value))
    #print(max_row_pre)
    for i in range(2, max_row_pre+1):
        relatType=nodeSheet['E{}'.format(i)].value
        startNode=findNode('process',nodeSheet['D{}'.format(i)].value)
        endNode=findNode('process',nodeSheet['F{}'.format(i)].value)
        r2=creatRelat(startNode,relatType,endNode)
        #r2=creatRelat(endNode,'后序工艺',startNode)#同时添加[后序工艺]联系
        print("{} creat".format(r2))
        #print("{} creat".format(r2))
    print("------created {} '前序工艺' relation------ ".format((max_row_pre-1))) 
    #初始加工
    max_row_init = max((bb.row for bb in nodeSheet['H'] if bb.value))
    #print(max_row_max_row_init)
    for i in range(2, max_row_init+1):
        relatType=nodeSheet['H{}'.format(i)].value
        startNode=findNode('process',nodeSheet['G{}'.format(i)].value)
        endNode=findNode('feature',nodeSheet['I{}'.format(i)].value)
        r3=creatRelat(startNode,relatType,endNode)
        print("{} creat".format(r3))
    print("------created {} '初始加工' relation------ ".format(max_row_init-1)) 
    #可用设备
    max_row_equip = max((bb.row for bb in nodeSheet['K'] if bb.value))
    #print(max_row_max_row_init)
    for i in range(2, max_row_equip+1):
        relatType=nodeSheet['K{}'.format(i)].value
        startNode=findNode('process',nodeSheet['J{}'.format(i)].value)
        endNode=findNode('equipment',nodeSheet['L{}'.format(i)].value)
        r4=creatRelat(startNode,relatType,endNode)
        print("{} creat".format(r4))
    print("------created {} '可用设备' relation------ ".format(max_row_equip-1)) 
    #装备设备
    max_row_tool = max((bb.row for bb in nodeSheet['N'] if bb.value))
    for i in range(2, max_row_tool+1):
        relatType=nodeSheet['N{}'.format(i)].value
        startNode=findNode('equipment',nodeSheet['M{}'.format(i)].value)
        endNode=findNode('tool',nodeSheet['O{}'.format(i)].value)
        r5=creatRelat(startNode,relatType,endNode)
        print("{} creat".format(r5))
    print("------created {} '装备刀具' relation------ ".format(max_row_tool-1)) 

def nodes2xlsx():
    workbook = load_workbook(".\data2.xlsx")
    #print(workbook.sheetnames)
    nodeSheet = workbook['nodeData']
    #特征节点
    nodes = list(matcher.match("feature"))
    max_row_feature=len(nodes)
    for i in range(2, max_row_feature+2): 
        nodeSheet.cell(row=i,column=1).value="feature"
        nodeSheet.cell(row=i,column=2).value=nodes[i-2]["name"]
    #工艺节点
    nodes = list(matcher.match("process"))
    max_row_process=len(nodes)
    for i in range(2, max_row_process+2): 
        nodeSheet.cell(row=i,column=3).value="process"
        nodeSheet.cell(row=i,column=4).value=nodes[i-2]["name"]
        nodeSheet.cell(row=i,column=5).value=nodes[i-2]["lowerIT"]
        nodeSheet.cell(row=i,column=6).value=nodes[i-2]["upperIT"]
        nodeSheet.cell(row=i,column=7).value=nodes[i-2]["lowerRa"]
        nodeSheet.cell(row=i,column=8).value=nodes[i-2]["upperRa"]
    #设备节点
    nodes = list(matcher.match("equipment"))
    max_row_equipment=len(nodes)
    for i in range(2, max_row_equipment+2): 
        nodeSheet.cell(row=i,column=9).value="equipment"
        nodeSheet.cell(row=i,column=10).value=nodes[i-2]["name"]
        nodeSheet.cell(row=i,column=11).value=nodes[i-2]["lowerSize"]
        nodeSheet.cell(row=i,column=12).value=nodes[i-2]["upperSize"]
        nodeSheet.cell(row=i,column=13).value=nodes[i-2]["lowerIT"]
        nodeSheet.cell(row=i,column=14).value=nodes[i-2]["upperIT"]
        nodeSheet.cell(row=i,column=15).value=nodes[i-2]["notes"]
    workbook.save(".\data2.xlsx")

def relats2xlsx():   
    workbook = load_workbook(".\data2.xlsx")
    #print(workbook.sheetnames)
    nodeSheet = workbook['relatData']
    #可用工艺
    relats=list(rmatcher.match(r_type='可用工艺'))
    max_row=len(relats)
    for i in range(2, max_row+2): 
        nodeSheet.cell(row=i,column=1).value=relats[i-2].start_node['name']
        nodeSheet.cell(row=i,column=2).value='可用工艺'
        nodeSheet.cell(row=i,column=3).value=relats[i-2].end_node['name']
    #前序工艺
    relats=list(rmatcher.match(r_type='前序工艺'))
    max_row=len(relats)
    for i in range(2, max_row+2): 
        nodeSheet.cell(row=i,column=4).value=relats[i-2].start_node['name']
        nodeSheet.cell(row=i,column=5).value='前序工艺'
        nodeSheet.cell(row=i,column=6).value=relats[i-2].end_node['name']
    #初始加工
    relats=list(rmatcher.match(r_type='初始加工'))
    max_row=len(relats)
    for i in range(2, max_row+2): 
        nodeSheet.cell(row=i,column=7).value=relats[i-2].start_node['name']
        nodeSheet.cell(row=i,column=8).value='初始加工'
        nodeSheet.cell(row=i,column=9).value=relats[i-2].end_node['name']
    #可用设备
    relats=list(rmatcher.match(r_type='可用设备'))
    max_row=len(relats)
    for i in range(2, max_row+2): 
        nodeSheet.cell(row=i,column=10).value=relats[i-2].start_node['name']
        nodeSheet.cell(row=i,column=11).value='可用设备'
        nodeSheet.cell(row=i,column=12).value=relats[i-2].end_node['name']
    workbook.save(".\data2.xlsx")

def printpath(path):
    nodes=path.nodes
    relats = path.relationships
    nodesNum=len(nodes)
    #relatsNum=len(relats)
    processList=[]
    bestIT=nodes[0]['lowerIT']
    bestRa=nodes[0]['lowerRa']
    for n in range(nodesNum-1):
        processList.append(nodes[n]['name'])
    processList.reverse()
    print("->".join(processList[i] for i in range(nodesNum-1)),end="")
    print("(最优结果:IT={} Ra={})".format(bestIT,bestRa))

def searchProcess(featureName,IT,Ra):
    print("---加工特征 [{}] 加工精度要求IT{} 粗糙度要求Ra{}---".format(featureName,IT,Ra))
    featureNode=findNode('feature',featureName)
    usableRelat=list(rmatcher.match([featureNode],r_type='可用工艺'))
    usableNum=len(usableRelat)
    #搜索符合特征要求的工艺
    indexlist=[]
    for i in range(usableNum):
        pname=usableRelat[i].end_node['name']
        bestIT=usableRelat[i].end_node['lowerIT']
        bestRa=usableRelat[i].end_node['lowerRa']
        if (IT>=bestIT)&(Ra>=bestRa):
            indexlist.append(i)
            #print("可用工艺[{}] 最优加工精度IT{} 表面粗糙度Ra{}".format(pname,bestIT,bestRa))
    #print(indexlist)
    finalNum=len(indexlist)
    pathNum=0
    '''
    def findpreprocess(processNode):
        r=list(rmatcher.match([processNode],r_type='前序工艺'))
        if r==None:
            return
        for i in r:
            if usableRelat.count()
    '''
    for i in range(finalNum):
        finalProcessNode=usableRelat[indexlist[i]].end_node
        #print(finalProcessNode['name'])
        
        '''
        cypher_ = "MATCH path=(m:process)-[:前序工艺|初始加工*1..8]->(n:feature) \
        WHERE m.name='研磨'AND n.name='平面' \
            RETURN path"
        '''
        #获取从末端工序到特征的路径
        cypher_ = "MATCH path=(m:process)-[:前序工艺|初始加工*1..8]->(n:feature) \
        WHERE m.name='{}'AND n.name='{}' \
            RETURN path".format(finalProcessNode['name'],featureNode['name'])
        s=graph.run(cypher_).to_series()
        sNum=len(s)#该末端工序到特征的路径数量
        #print(sNum)
        for j in range(sNum):
            #读取末端节点的前一个节点
            nodes=s[j].nodes
            preNode=nodes[1]
            preIT=preNode['lowerIT']
            preRa=preNode['lowerRa']
            if (IT>=preIT)&(Ra>=preRa):#若前一道工序满足加工要求，舍弃该路线
                continue
            else:
                pathNum=pathNum+1
                #print("path{}:{}".format(pathNum,s[j]))
                print("path{}:".format(pathNum))
                printpath(s[j])#输出正向加工路线
def buildGraph():
    graph.delete_all()
    xlsx2nodes()
    xlsx2relats()

if __name__ == "__main__":
    #buildGraph()
    searchProcess('孔',7,1.5)
    #searchProcess('平面',6,0.08)